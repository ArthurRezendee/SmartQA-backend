from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
from decimal import Decimal
import datetime



class ExcelService:

    def generate_excel(self, data: Any, sheet_name: str = "Dados") -> BytesIO:
        workbook = Workbook()

        # Remove aba padrão se for multi-sheet
        if isinstance(data, dict):
            workbook.remove(workbook.active)

            for name, rows in data.items():
                self._write_sheet(workbook, name, rows)

        else:
            self._write_sheet(workbook, sheet_name, data)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    # =============================
    # Conversão segura de valores
    # =============================

    def _sanitize_value(self, value):

        if value is None:
            return None

        if isinstance(value, (str, int, float, bool)):
            return value

        if isinstance(value, (datetime.date, datetime.datetime)):
            return value.isoformat()

        if isinstance(value, Decimal):
            return float(value)

        if isinstance(value, list):
            return json.dumps(
                [self._convert_complex(v) for v in value],
                ensure_ascii=False
            )

        if isinstance(value, dict):
            return json.dumps(
                {k: self._convert_complex(v) for k, v in value.items()},
                ensure_ascii=False
            )

        # Qualquer objeto ORM vira string
        return str(value)


    def _convert_complex(self, value):
        """
        Converte qualquer objeto complexo para algo serializável.
        """

        if isinstance(value, (str, int, float, bool)):
            return value

        if isinstance(value, (datetime.date, datetime.datetime)):
            return value.isoformat()

        if isinstance(value, Decimal):
            return float(value)

        if isinstance(value, dict):
            return {k: self._convert_complex(v) for k, v in value.items()}

        if isinstance(value, list):
            return [self._convert_complex(v) for v in value]

        # ORM object
        if hasattr(value, "__dict__"):
            return {
                k: self._convert_complex(v)
                for k, v in vars(value).items()
                if not k.startswith("_")
            }

        return str(value)

    def _object_to_dict(self, obj: Any) -> Dict:
        """
        Converte objeto ORM em dict ignorando atributos privados.
        """
        return {
            key: value
            for key, value in vars(obj).items()
            if not key.startswith("_")
        }
        
    def _write_sheet(self, workbook: Workbook, sheet_name: str, data: List[Any]):
        if not data:
            return

        if not isinstance(data[0], dict):
            data = [self._object_to_dict(obj) for obj in data]

        sheet = workbook.create_sheet(title=sheet_name)

        headers = list(data[0].keys())

        # Header
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = self._sanitize_value(row_data.get(header))
                sheet.cell(row=row_num, column=col_num, value=value)

        # Column width
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = max(15, len(header) + 2)